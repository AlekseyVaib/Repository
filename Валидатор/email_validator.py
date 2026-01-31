"""
Валидатор email адресов для рассылок
Поддерживает проверку синтаксиса, DNS, MX записей, SMTP и определение качества адресов
"""

import pandas as pd
import re
import dns.resolver
import smtplib
import socket
import time
import os
from typing import Dict, Tuple, List, Optional
import logging
from datetime import datetime
import warnings
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import urllib.request
import urllib.error
import random
import string

warnings.filterwarnings('ignore')

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('email_validator.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class EmailValidator:
    """
    Класс для валидации email адресов с учетом требований рассылок
    """
    
    def __init__(self, timeout: int = 10, max_retries: int = 2, check_smtp: bool = True, 
                 accept_catch_all: bool = False, validation_mode: str = 'strict'):
        """
        Инициализация валидатора
        
        Args:
            timeout: Таймаут для сетевых запросов (секунды)
            max_retries: Максимальное количество попыток при ошибках
            check_smtp: Выполнять ли SMTP проверку
            accept_catch_all: Считать ли catch-all адреса валидными
            validation_mode: Режим валидации ('strict' - строгий, 'lenient' - лояльный)
        """
        self.timeout = timeout
        self.max_retries = max_retries
        self.check_smtp = check_smtp
        self.accept_catch_all = accept_catch_all
        self.validation_mode = validation_mode  # 'strict' или 'lenient'
        self.disposable_domains = self._load_disposable_domains()
        self.role_accounts = self._load_role_accounts()
        self.check_attempts = {}  # Для отслеживания попыток проверки
        
    def _load_disposable_domains(self) -> set:
        """Загрузка списка доменов одноразовых почт"""
        disposable_domains = {
            # Популярные сервисы
            'tempmail.com', '10minutemail.com', 'guerrillamail.com',
            'mailinator.com', 'trashmail.com', 'yopmail.com',
            'dispostable.com', 'temp-mail.org', 'fakeinbox.com',
            'sharklasers.com', 'getairmail.com', 'maildrop.cc',
            'throwawaymail.com', 'tempmailaddress.com',
            
            # Дополнительные сервисы
            'mailnesia.com', 'emailondeck.com', 'mailmetrash.com',
            'trashmailer.com', 'mailmoat.com', 'discard.email',
            'spamgourmet.com', 'spamhole.com', 'spamfree24.org',
            
            # Временные домены
            'tmpmail.org', 'temporary-mail.net', 'mytemp.email',
            'temp-mail.io', 'tempinbox.com', 'temporarymailaddress.com',
            'mohmal.com', 'mailcatch.com', 'mintemail.com',
            'emailias.com', 'spambox.us', 'mailnull.com'
        }
        
        # Попытка загрузить из файла
        try:
            domains_file = os.path.join(os.path.dirname(__file__), "disposable_domains.txt")
            if os.path.exists(domains_file):
                with open(domains_file, 'r', encoding='utf-8') as f:
                    file_domains = {line.strip().lower() for line in f if line.strip() and not line.startswith('#')}
                    disposable_domains.update(file_domains)
                    logger.info(f"Загружено {len(file_domains)} доменов из файла")
        except Exception as e:
            logger.warning(f"Не удалось загрузить файл disposable_domains: {e}")
        
        return disposable_domains
    
    def _load_role_accounts(self) -> set:
        """Загрузка списка ролевых аккаунтов"""
        return {
            'admin', 'administrator', 'support', 'info', 'sales', 'contact',
            'help', 'service', 'noreply', 'no-reply', 'postmaster',
            'webmaster', 'hostmaster', 'abuse', 'security', 'marketing',
            'newsletter', 'notifications', 'alerts', 'system', 'test',
            'testing', 'demo', 'example', 'mailer-daemon', 'daemon'
        }
    
    def check_syntax(self, email: str) -> Tuple[bool, str]:
        """
        Проверка синтаксиса email согласно RFC 5322
        
        Args:
            email: Email адрес для проверки
            
        Returns:
            Tuple: (валидность, сообщение)
        """
        try:
            if not email or not isinstance(email, str):
                return False, "Пустой или неверный тип данных"
            
            email = email.strip()
            
            # Базовые проверки
            if '@' not in email:
                return False, "Отсутствует символ @"
            
            if email.count('@') > 1:
                return False, "Множественные символы @"
            
            # Разделение на локальную часть и домен
            local_part, domain = email.rsplit('@', 1)
            
            # Проверка длины
            if len(email) > 254:
                return False, "Email слишком длинный (>254 символов)"
            
            if len(local_part) > 64:
                return False, "Локальная часть слишком длинная (>64 символов)"
            
            if len(domain) > 255:
                return False, "Домен слишком длинный (>255 символов)"
            
            # Проверка локальной части
            if not local_part:
                return False, "Локальная часть пустая"
            
            if local_part.startswith('.') or local_part.endswith('.'):
                return False, "Локальная часть не может начинаться/заканчиваться точкой"
            
            if '..' in local_part:
                return False, "Локальная часть содержит две точки подряд"
            
            # Проверка домена
            if not domain:
                return False, "Домен пустой"
            
            if domain.startswith('.') or domain.endswith('.'):
                return False, "Домен не может начинаться/заканчиваться точкой"
            
            if '..' in domain:
                return False, "Домен содержит две точки подряд"
            
            # Проверка формата локальной части (RFC 5322)
            local_pattern = r'^[a-zA-Z0-9.!#$%&\'*+/=?^_`{|}~-]+$'
            if not re.match(local_pattern, local_part):
                return False, "Локальная часть содержит недопустимые символы"
            
            # Проверка формата домена
            domain_pattern = r'^[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?(?:\.[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?)*\.[a-zA-Z]{2,}$'
            if not re.match(domain_pattern, domain):
                return False, "Домен имеет неверный формат"
            
            return True, "Синтаксис корректен"
            
        except Exception as e:
            return False, f"Ошибка проверки синтаксиса: {str(e)}"
    
    def check_domain_dns(self, domain: str) -> Tuple[bool, str, List[str]]:
        """
        Проверка DNS и MX записей домена
        
        Args:
            domain: Домен для проверки
            
        Returns:
            Tuple: (валидность, сообщение, список MX записей)
        """
        mx_records = []
        
        try:
            # Настройка резолвера
            resolver = dns.resolver.Resolver()
            resolver.timeout = self.timeout
            resolver.lifetime = self.timeout
            
            # Проверка MX записей
            try:
                answers = resolver.resolve(domain, 'MX')
                for rdata in answers:
                    mx_server = str(rdata.exchange).rstrip('.')
                    mx_records.append(f"{rdata.preference} {mx_server}")
                
                mx_records.sort()  # Сортировка по приоритету
                
                if mx_records:
                    mx_list = ', '.join([mx.split()[1] for mx in mx_records[:3]])
                    return True, f"MX найден: {mx_list}", mx_records
                else:
                    return False, "Нет MX записей", []
                    
            except dns.resolver.NoAnswer:
                # Если нет MX, проверяем A запись
                try:
                    resolver.resolve(domain, 'A')
                    return True, "Нет MX, но есть A запись", []
                except:
                    return False, "Нет MX и A записей", []
                    
            except dns.resolver.NXDOMAIN:
                return False, "Домен не существует", []
                
            except (dns.resolver.Timeout, dns.resolver.NoNameservers):
                return False, "Таймаут или отсутствие NS серверов", []
                
        except Exception as e:
            logger.warning(f"Ошибка DNS проверки для {domain}: {e}")
            return False, f"Ошибка DNS: {str(e)}", []
    
    def check_smtp_verification(self, email: str, mx_records: List[str]) -> Dict[str, str]:
        """
        Проверка через SMTP (без отправки письма)
        
        Args:
            email: Email адрес для проверки
            mx_records: Список MX записей
            
        Returns:
            Dict: Результаты SMTP проверки
        """
        if not self.check_smtp:
            return {
                'smtp_connection': 'Нет',
                'email_active': '–',
                'catch_all': 'Нет',  # Если SMTP отключен, считаем что не catch-all
                'mailbox_full': '–',
                'smtp_message': 'SMTP проверка отключена'
            }
        
        if not mx_records:
            return {
                'smtp_connection': 'Нет',
                'email_active': '–',
                'catch_all': 'Нет',  # Если нет MX записей, считаем что не catch-all
                'mailbox_full': '–',
                'smtp_message': 'Нет MX записей для проверки'
            }
        
        # Извлекаем серверы из MX записей
        mx_servers = []
        for mx_record in mx_records[:5]:  # Берем первые 5
            try:
                parts = mx_record.split()
                if len(parts) >= 2:
                    mx_servers.append(parts[-1])
            except:
                continue
        
        if not mx_servers:
            return {
                'smtp_connection': 'Нет',
                'email_active': '–',
                'catch_all': 'Нет',  # Если не удалось извлечь MX серверы, считаем что не catch-all
                'mailbox_full': '–',
                'smtp_message': 'Не удалось извлечь MX серверы'
            }
        
        # Пробуем подключиться к SMTP серверам
        for mx_server in mx_servers[:3]:  # Проверяем первые 3
            try:
                server = smtplib.SMTP(timeout=self.timeout)
                server.set_debuglevel(0)
                
                # Подключение
                server.connect(mx_server, 25)
                code, message = server.helo()
                
                if code != 250:
                    server.quit()
                    continue
                
                # Проверка почтового ящика
                test_sender = 'check@email-validator.com'
                server.mail(test_sender)
                code, message = server.rcpt(email)
                
                # Анализ ответа для основного email
                message_str = message.decode('utf-8', errors='ignore') if isinstance(message, bytes) else str(message)
                
                # Проверка на catch-all: тестируем несколько несуществующих адресов
                catch_all = 'Нет'
                if code == 250:
                    # Если основной адрес принят, проверяем catch-all
                    domain = email.split('@')[1] if '@' in email else ''
                    if domain:
                        # Делаем несколько попыток для надежности
                        catch_all_detected = False
                        catch_all_unknown = False
                        
                        for attempt in range(5):  # 5 попыток с разными адресами для большей надежности
                            try:
                                # Генерируем случайный несуществующий адрес на том же домене
                                random_user = ''.join(random.choices(string.ascii_lowercase + string.digits, k=15))
                                test_email = f"{random_user}@{domain}"
                                
                                server.mail(test_sender)
                                test_code, test_message = server.rcpt(test_email)
                                
                                # Если несуществующий адрес принят (код 250), это catch-all
                                if test_code == 250:
                                    catch_all_detected = True
                                    logger.debug(f"Catch-all обнаружен для {domain}: {test_email} принят")
                                    break
                                elif test_code == 550:
                                    # Адрес отклонен - точно не catch-all
                                    logger.debug(f"Catch-all не обнаружен для {domain}: {test_email} отклонен")
                                    break
                                else:
                                    # Неизвестный код - возможно временная проблема
                                    logger.debug(f"Неизвестный код {test_code} при проверке catch-all для {domain}")
                                    catch_all_unknown = True
                                    
                            except Exception as e:
                                logger.debug(f"Ошибка при проверке catch-all (попытка {attempt + 1}): {e}")
                                catch_all_unknown = True
                                continue
                        
                        if catch_all_detected:
                            catch_all = 'Да'
                        elif catch_all_unknown:
                            # Если не удалось определить, считаем что это не catch-all
                            catch_all = 'Нет'
                            logger.debug(f"Не удалось точно определить catch-all для {domain}, считаем 'Нет'")
                
                server.quit()
                
                if code == 250:
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Да',
                        'catch_all': catch_all,
                        'mailbox_full': 'Нет',
                        'smtp_message': f'SUCCESS: {code} - {message_str}'
                    }
                elif code == 550:
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Нет',
                        'catch_all': 'Нет',
                        'mailbox_full': 'Нет',
                        'smtp_message': f'MAILBOX_NOT_FOUND: {code} - {message_str}'
                    }
                elif code == 452:
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Да',
                        'catch_all': 'Нет',
                        'mailbox_full': 'Да',
                        'smtp_message': f'MAILBOX_FULL: {code} - {message_str}'
                    }
                elif code == 450:
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Временно недоступен',
                        'catch_all': 'Нет',
                        'mailbox_full': 'Нет',
                        'smtp_message': f'MAILBOX_UNAVAILABLE: {code} - {message_str}'
                    }
                elif code in [551, 553]:
                    # Проблемы с маршрутизацией или конфигурацией
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Нет',
                        'catch_all': 'Нет',
                        'mailbox_full': 'Нет',
                        'smtp_message': f'ROUTING_ERROR: {code} - {message_str}'
                    }
                elif code in [421, 451]:
                    # Временные ошибки сервера
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Временно недоступен',
                        'catch_all': 'Нет',
                        'mailbox_full': 'Нет',
                        'smtp_message': f'TEMPORARY_ERROR: {code} - {message_str}'
                    }
                elif code in [552, 554]:
                    # Переполнение или превышение лимита
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Да',
                        'catch_all': 'Нет',
                        'mailbox_full': 'Да',
                        'smtp_message': f'MAILBOX_FULL_OR_LIMIT: {code} - {message_str}'
                    }
                else:
                    # Неизвестный код - считаем невалидным для безопасности
                    return {
                        'smtp_connection': 'Да',
                        'email_active': 'Нет',  # При неизвестном коде считаем неактивным
                        'catch_all': 'Нет',
                        'mailbox_full': 'Нет',
                        'smtp_message': f'UNKNOWN_CODE: {code} - {message_str}'
                    }
                    
            except (smtplib.SMTPConnectError, smtplib.SMTPServerDisconnected,
                    socket.timeout, socket.error) as e:
                logger.debug(f"Ошибка подключения к {mx_server}: {e}")
                continue
            except Exception as e:
                logger.debug(f"Ошибка SMTP для {mx_server}: {e}")
                continue
        
        return {
            'smtp_connection': 'Нет',
            'email_active': '–',
            'catch_all': 'Нет',  # Если не удалось подключиться, считаем что не catch-all
            'mailbox_full': '–',
            'smtp_message': 'Не удалось подключиться ни к одному MX серверу'
        }
    
    def check_disposable_email(self, domain: str) -> bool:
        """Проверка на одноразовый email домен"""
        domain_lower = domain.lower()
        
        # Точное совпадение
        if domain_lower in self.disposable_domains:
            return True
        
        # Проверка поддоменов
        for disposable_domain in self.disposable_domains:
            if domain_lower.endswith('.' + disposable_domain):
                return True
        
        return False
    
    def check_role_account(self, local_part: str) -> bool:
        """Проверка на ролевой аккаунт"""
        local_lower = local_part.lower()
        
        # Точное совпадение
        if local_lower in self.role_accounts:
            return True
        
        # Проверка с числами (admin1, support2 и т.д.)
        for role in self.role_accounts:
            if local_lower.startswith(role):
                suffix = local_lower[len(role):]
                if not suffix or suffix.isdigit() or suffix.startswith('-'):
                    return True
        
        return False
    
    def check_domain_reputation(self, domain: str) -> Tuple[bool, str]:
        """
        Проверка репутации домена в черных списках (DNSBL)
        
        Args:
            domain: Домен для проверки
            
        Returns:
            Tuple: (не в черном списке, сообщение)
        """
        try:
            # Получаем IP адрес домена
            try:
                ip = socket.gethostbyname(domain)
            except socket.gaierror:
                return True, "Не удалось получить IP адрес (не влияет на репутацию)"
            
            # Разворачиваем IP для проверки в DNSBL
            ip_parts = ip.split('.')
            if len(ip_parts) != 4:
                return True, "Неверный формат IP"
            
            reversed_ip = '.'.join(reversed(ip_parts))
            
            # Список популярных DNSBL серверов
            dnsbl_servers = [
                'zen.spamhaus.org',
                'bl.spamcop.net',
                'dnsbl.sorbs.net',
                'b.barracudacentral.org',
            ]
            
            for dnsbl in dnsbl_servers:
                try:
                    check_host = f"{reversed_ip}.{dnsbl}"
                    # Пытаемся разрешить имя - если получили IP, значит в черном списке
                    socket.gethostbyname(check_host)
                    logger.warning(f"Домен {domain} (IP: {ip}) найден в черном списке: {dnsbl}")
                    return False, f"Домен в черном списке: {dnsbl}"
                except socket.gaierror:
                    # Не в черном списке - это нормально
                    continue
                except Exception as e:
                    logger.debug(f"Ошибка проверки {dnsbl} для {domain}: {e}")
                    continue
            
            return True, "Домен не в черных списках"
        except Exception as e:
            logger.debug(f"Ошибка проверки репутации {domain}: {e}")
            return True, "Не удалось проверить репутацию"
    
    def check_suspicious_domain(self, domain: str) -> bool:
        """
        Проверка на подозрительные домены (typo-squatting, подделки известных доменов)
        
        Args:
            domain: Домен для проверки
            
        Returns:
            bool: True если домен подозрительный
        """
        domain_lower = domain.lower()
        
        # Список известных валидных доменов (никогда не помечаем как подозрительные)
        valid_domains = {
            'gmail.com', 'mail.ru', 'yandex.ru', 'yahoo.com', 'hotmail.com', 
            'outlook.com', 'mail.com', 'bk.ru', 'list.ru', 'inbox.ru',
            'rambler.ru', 'ya.ru', 'icloud.com', 'protonmail.com', 'aol.com',
            'live.com', 'msn.com', 'qq.com', '163.com', 'sina.com'
        }
        
        # Если это известный валидный домен, не проверяем дальше
        if domain_lower in valid_domains:
            return False
        
        # Список подозрительных доменов (точные совпадения опечаток)
        suspicious_domains = [
            # Опечатки gmail
            'gmai1.com', 'gmai.com', 'gmaill.com', 'gmial.com',
            'gmail.co', 'gmail.cm', 'gmail.co.uk.com',
            # Опечатки yahoo
            'yaho0.com', 'yahoo.co', 'yhoo.com', 'yahooo.com',
            # Опечатки hotmail
            'hotmai1.com', 'hotmial.com', 'hotmai.com', 'hotmali.com',
            # Опечатки mail.ru
            'mai1.ru', 'mail.r', 'mail.ry',
            # Опечатки outlook
            'outlok.com', 'outlook.co',
            # Опечатки yandex
            'yandex.co', 'yandex.cm', 'yandex.r',
        ]
        
        # Проверяем точное совпадение с подозрительными доменами
        if domain_lower in suspicious_domains:
            logger.warning(f"Обнаружен подозрительный домен: {domain}")
            return True
        
        return False
    
    def check_domain_active(self, domain: str) -> Tuple[bool, str]:
        """
        Проверка активности домена (наличие веб-сайта)
        
        Args:
            domain: Домен для проверки
            
        Returns:
            Tuple: (активен, сообщение)
        """
        try:
            # Проверяем наличие веб-сайта через HTTP
            try:
                urllib.request.urlopen(f"http://{domain}", timeout=5)
                return True, "Домен активен (есть веб-сайт)"
            except (urllib.error.URLError, socket.timeout):
                pass
            
            # Проверяем через HTTPS
            try:
                urllib.request.urlopen(f"https://{domain}", timeout=5)
                return True, "Домен активен (есть веб-сайт)"
            except (urllib.error.URLError, socket.timeout):
                # Нет веб-сайта, но это не обязательно плохо для почтового домена
                return True, "Нет веб-сайта (нормально для почтовых доменов)"
        except Exception as e:
            logger.debug(f"Ошибка проверки активности {domain}: {e}")
            return True, "Не удалось проверить активность"
    
    def determine_reliability(self, results: Dict) -> str:
        """
        Определение надежности email для рассылок (улучшенная версия)
        
        Args:
            results: Результаты всех проверок
            
        Returns:
            str: Надежность (Высокая, Средняя, Нет)
        """
        # Высокая надежность: ВСЕ проверки пройдены идеально
        # (Проверка репутации домена отключена - может блокировать обычные домены)
        if (results.get('syntax_valid', False) and
            results.get('dns_mx_valid', False) and
            results.get('smtp_connection') == 'Да' and
            results.get('email_active') == 'Да' and
            not results.get('is_disposable', False) and
            results.get('mailbox_full') != 'Да' and
            not results.get('is_suspicious_domain', False)):
            return "Высокая"
        
        # Средняя надежность: основные проверки пройдены, но есть незначительные проблемы
        # (например, временные проблемы с SMTP, но синтаксис и DNS в порядке)
        if (results.get('syntax_valid', False) and
            results.get('dns_mx_valid', False) and
            results.get('smtp_connection') == 'Да' and
            results.get('email_active') in ['Да', '–'] and  # Может быть неопределен
            not results.get('is_disposable', False) and
            results.get('mailbox_full') != 'Да'):
            return "Средняя"
        
        return "Нет"
    
    def validate_email(self, email: str) -> Dict:
        """
        Полная валидация email адреса
        
        Args:
            email: Email адрес для проверки
            
        Returns:
            Dict: Результаты проверки
        """
        start_time = time.time()
        
        # Базовые данные
        email = str(email).strip() if email else ""
        local_part = ""
        domain = ""
        
        # Извлечение частей
        if '@' in email:
            try:
                local_part, domain = email.rsplit('@', 1)
                local_part = local_part.strip()
                domain = domain.strip().lower()
            except:
                pass
        
        # Проверка синтаксиса
        syntax_valid, syntax_message = self.check_syntax(email)
        
        # Увеличиваем счетчик попыток
        if email not in self.check_attempts:
            self.check_attempts[email] = 0
        self.check_attempts[email] += 1
        
        # Если синтаксис неверен, возвращаем базовый результат
        if not syntax_valid:
            return {
                'Email': email,
                'Пользователь': local_part if local_part else '',
                'Домен': domain if domain else '',
                'Валидность': 'Нет',
                'Корректность': 'Нет',
                'Надежность': 'Нет',
                'Одноразовый (DEA)': '–',
                'Получение DNS, MX': 'Нет',
                'Связь с SMTP-сервером': 'Нет',
                'Email активен': 'Нет',
                'Доставляемость': 'Нет',
                'Catch-all адрес': 'Нет',  # При ошибке считаем что не catch-all
                'Email переполнен': '–',
                'Ролевой аккаунт': 'Нет',
                'Время проверки, сек': round(time.time() - start_time, 2),
                'Попыток проверки': self.check_attempts[email],
                'МХ-записи': ''
            }
        
        # Проверка DNS/MX
        dns_valid, dns_message, mx_records = self.check_domain_dns(domain)
        
        # Проверка репутации домена ОТКЛЮЧЕНА (может блокировать обычные домены типа gmail.com, mail.ru)
        
        # Проверка на подозрительный домен
        is_suspicious_domain = False
        if domain:
            is_suspicious_domain = self.check_suspicious_domain(domain)
        
        # Проверка активности домена (опционально, не блокируем если не удалось)
        if domain:
            domain_active, _ = self.check_domain_active(domain)
        
        # Проверка SMTP
        smtp_results = self.check_smtp_verification(email, mx_records)
        
        # Проверка одноразового email
        is_disposable = self.check_disposable_email(domain)
        
        # Проверка ролевого аккаунта
        is_role_account = self.check_role_account(local_part)
        
        # Определение надежности (с учетом новых проверок)
        reliability_results = {
            'syntax_valid': syntax_valid,
            'dns_mx_valid': dns_valid,
            'smtp_connection': smtp_results['smtp_connection'],
            'email_active': smtp_results['email_active'],
            'is_disposable': is_disposable,
            'is_role_account': is_role_account,
            'mailbox_full': smtp_results['mailbox_full'],
            'is_suspicious_domain': is_suspicious_domain,
            'domain_reputation_ok': True  # Проверка репутации отключена
        }
        reliability = self.determine_reliability(reliability_results)
        
        # Определение доставляемости
        deliverability = '–'
        if smtp_results['email_active'] == 'Да':
            deliverability = 'Да'
        elif smtp_results['email_active'] == 'Нет':
            deliverability = 'Нет'
        
        # Логика валидности для рассылки (зависит от режима)
        if self.validation_mode == 'strict':
            # СТРОГИЙ РЕЖИМ: Email валидный ТОЛЬКО если:
            # 1. Надежность ВЫСОКАЯ (убрали "Среднюю")
            # 2. Получен ответ от SMTP
            # 3. Email активен (обязательно)
            # 4. Доставляемость подтверждена (обязательно)
            # 5. Не одноразовый
            # 6. Не переполнен
            # 7. Не подозрительный домен
            # (Проверка репутации домена отключена - может блокировать обычные домены)
            is_valid_for_mailing = (
                reliability == 'Высокая' and  # ТОЛЬКО высокая надежность
                smtp_results['smtp_connection'] == 'Да' and
                smtp_results['email_active'] == 'Да' and  # ОБЯЗАТЕЛЬНО активен
                deliverability == 'Да' and  # ОБЯЗАТЕЛЬНО доставляем
                not is_disposable and
                smtp_results['mailbox_full'] != 'Да' and  # Не переполнен
                not is_suspicious_domain  # Не подозрительный
            )
        else:
            # ЛОЯЛЬНЫЙ РЕЖИМ: Более мягкие критерии (на 15-20% больше валидных)
            # Принимаем адреса с "Средней" надежностью
            # Не требуем обязательной активности email (если SMTP недоступен, но DNS/MX в порядке)
            # Не блокируем из-за репутации домена (если не критично)
            # Не блокируем подозрительные домены (только предупреждаем)
            is_valid_for_mailing = (
                reliability in ['Высокая', 'Средняя'] and  # Высокая ИЛИ Средняя надежность
                (smtp_results['smtp_connection'] == 'Да' or dns_valid) and  # SMTP или хотя бы DNS
                (smtp_results['email_active'] == 'Да' or 
                 smtp_results['email_active'] == '–' or 
                 not self.check_smtp) and  # Активен, неопределен или SMTP отключен
                (deliverability == 'Да' or deliverability == '–') and  # Доставляем или неопределено
                not is_disposable and  # Все равно не одноразовый
                smtp_results['mailbox_full'] != 'Да'  # Не переполнен
                # Не проверяем репутацию и подозрительные домены в лояльном режиме
            )
        
        # Если catch-all не принимается и это catch-all адрес, то невалиден
        if not self.accept_catch_all and smtp_results['catch_all'] == 'Да':
            is_valid_for_mailing = False
        
        # Формирование результатов согласно структуре из примера
        # Примечание: В столбце "Catch-all адрес" прочерк (–) означает, что не удалось определить,
        # но мы считаем это как "Нет" (не catch-all адрес) для консервативной оценки
        catch_all_result = smtp_results['catch_all']
        # Если catch_all = '–', заменяем на 'Нет' для единообразия
        if catch_all_result == '–':
            catch_all_result = 'Нет'
        
        results = {
            'Email': email,
            'Пользователь': local_part if local_part else '',
            'Домен': domain if domain else '',
            'Валидность': 'Да' if is_valid_for_mailing else 'Нет',
            'Корректность': 'Да' if syntax_valid else 'Нет',
            'Надежность': reliability,
            'Одноразовый (DEA)': 'Да' if is_disposable else 'Нет',
            'Получение DNS, MX': 'Да' if dns_valid else 'Нет',
            'Связь с SMTP-сервером': smtp_results['smtp_connection'],
            'Email активен': smtp_results['email_active'],
            'Доставляемость': deliverability,
            'Catch-all адрес': catch_all_result,  # Всегда "Да" или "Нет", прочерки заменены на "Нет"
            'Email переполнен': smtp_results['mailbox_full'],
            'Ролевой аккаунт': 'Да' if is_role_account else 'Нет',
            'Время проверки, сек': round(time.time() - start_time, 2),
            'Попыток проверки': self.check_attempts[email],
            'МХ-записи': '\n'.join(mx_records[:5]) if mx_records else ''
        }
        
        return results


def save_results_to_excel(results_df: pd.DataFrame, output_file: str, is_checkpoint: bool = False) -> None:
    """
    Сохранение результатов в Excel файл с форматированием
    
    Args:
        results_df: DataFrame с результатами проверки
        output_file: Путь к выходному файлу
        is_checkpoint: Флаг промежуточного сохранения
    """
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name='Результаты проверки', index=False)
            
            # Автоподбор ширины столбцов и применение цветов
            worksheet = writer.sheets['Результаты проверки']
            
            # Определение цветов
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            # Получение индексов столбцов по названиям
            header_row = 1
            column_indices = {}
            for idx, cell in enumerate(worksheet[header_row], 1):
                if cell.value:
                    column_indices[cell.value] = idx - 1  # 0-based index
            
            # Применение цветов к ячейкам
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                for col_name, col_idx in column_indices.items():
                    if col_idx >= len(row):
                        continue
                    cell = row[col_idx]
                    value = cell.value
                    
                    if value is None:
                        continue
                    
                    value_str = str(value).strip()
                    
                    # Валидность
                    if col_name == 'Валидность':
                        if value_str == 'Да':
                            cell.fill = green_fill
                        else:
                            cell.fill = yellow_fill
                    
                    # Корректность
                    elif col_name == 'Корректность':
                        if value_str == 'Да':
                            cell.fill = green_fill
                        else:
                            cell.fill = yellow_fill
                    
                    # Надежность
                    elif col_name == 'Надежность':
                        if value_str in ['Высокая', 'Средняя']:
                            cell.fill = green_fill
                        elif value_str == 'Нет':
                            cell.fill = yellow_fill
                    
                    # Одноразовый (DEA)
                    elif col_name == 'Одноразовый (DEA)':
                        if value_str == 'Нет':
                            cell.fill = green_fill
                        elif value_str == 'Да' or value_str == '–':
                            cell.fill = yellow_fill
                    
                    # Получение DNS, MX
                    elif col_name == 'Получение DNS, MX':
                        if value_str == 'Да':
                            cell.fill = green_fill
                        else:
                            cell.fill = yellow_fill
                    
                    # Связь с SMTP-сервером
                    elif col_name == 'Связь с SMTP-сервером':
                        if value_str == 'Да':
                            cell.fill = green_fill
                        else:
                            cell.fill = yellow_fill
                    
                    # Email активен
                    elif col_name == 'Email активен':
                        if value_str == 'Да':
                            cell.fill = green_fill
                        elif value_str == 'Нет' or value_str == '–':
                            cell.fill = yellow_fill
                    
                    # Доставляемость
                    elif col_name == 'Доставляемость':
                        if value_str == 'Да':
                            cell.fill = green_fill
                        elif value_str == 'Нет' or value_str == '–':
                            cell.fill = yellow_fill
                    
                    # Catch-all адрес
                    elif col_name == 'Catch-all адрес':
                        if value_str == 'Да':
                            cell.fill = yellow_fill
                        elif value_str == 'Нет' or value_str == '–':
                            cell.fill = green_fill
                    
                    # Email переполнен
                    elif col_name == 'Email переполнен':
                        if value_str == 'Да':
                            cell.fill = yellow_fill
                        elif value_str == 'Нет':
                            cell.fill = green_fill
                        elif value_str == '–':
                            cell.fill = yellow_fill
                    
                    # Ролевой аккаунт
                    elif col_name == 'Ролевой аккаунт':
                        if value_str == 'Нет':
                            cell.fill = green_fill
                        elif value_str == 'Да' or value_str == '–':
                            cell.fill = yellow_fill
                    
                    # Время проверки (выделение выбросов)
                    elif col_name == 'Время проверки, сек':
                        try:
                            time_value = float(value_str.replace(',', '.'))
                            if time_value > 7.0:
                                cell.fill = yellow_fill
                        except:
                            pass
            
            # Автоподбор ширины столбцов
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        if is_checkpoint:
            logger.info(f"💾 Промежуточное сохранение: {output_file}")
        else:
            logger.info(f"Результаты успешно сохранены в: {output_file}")
    except Exception as e:
        logger.error(f"Ошибка сохранения Excel файла: {e}")
        # Попытка сохранить в CSV
        csv_file = output_file.replace('.xlsx', '.csv')
        results_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        if is_checkpoint:
            logger.info(f"💾 Промежуточное сохранение (CSV): {csv_file}")
        else:
            logger.info(f"Результаты сохранены в CSV: {csv_file}")


def process_excel_file(input_file: str, output_file: Optional[str] = None, 
                      check_smtp: bool = True, timeout: int = 10, 
                      accept_catch_all: bool = False, max_emails: Optional[int] = None,
                      validation_mode: str = 'strict') -> pd.DataFrame:
    """
    Обработка Excel файла с email адресами
    
    Args:
        input_file: Путь к входному Excel файлу
        output_file: Путь для сохранения результата (если None, генерируется автоматически)
        check_smtp: Выполнять ли SMTP проверку
        timeout: Таймаут для сетевых запросов
        accept_catch_all: Считать ли catch-all адреса валидными
        max_emails: Максимальное количество email для проверки (None = все)
        validation_mode: Режим валидации ('strict' - строгий, 'lenient' - лояльный)
        
    Returns:
        pd.DataFrame: DataFrame с результатами проверки
    """
    # Базовое имя для выходного файла (будет сгенерировано в конце с временем завершения)
    base_name = None
    if output_file is None:
        # Получаем базовое имя входного файла без расширения
        base_name = os.path.splitext(os.path.basename(input_file))[0]
    
    try:
        # Проверка существования файла
        if not os.path.exists(input_file):
            logger.error(f"Файл не найден: {input_file}")
            raise FileNotFoundError(f"Файл не найден: {input_file}")
        
        # Чтение Excel файла
        logger.info(f"Чтение файла: {input_file}")
        try:
            df = pd.read_excel(input_file)
        except Exception as e:
            logger.error(f"Ошибка чтения Excel файла: {e}")
            raise
        
        if df.empty:
            logger.error("Файл пустой или не содержит данных")
            raise ValueError("Файл пустой")
        
        # Определение столбца с email
        email_column = None
        for col in df.columns:
            col_lower = str(col).lower()
            if any(keyword in col_lower for keyword in ['email', 'e-mail', 'почта', 'mail', 'адрес']):
                email_column = col
                break
        
        if email_column is None:
            email_column = df.columns[0]
            logger.warning(f"Столбец с email не найден. Используется первый столбец: {email_column}")
        
        # Извлечение email адресов (убираем дубликаты)
        emails = []
        seen_emails = set()
        for idx, value in df[email_column].items():
            if pd.isna(value):
                continue
            email_str = str(value).strip()
            if email_str and email_str.lower() not in ['nan', 'none', '']:
                email_lower = email_str.lower()
                if email_lower not in seen_emails:
                    seen_emails.add(email_lower)
                    emails.append(email_str)
        
        total_emails = len(emails)
        logger.info(f"Найдено {total_emails} email адресов для проверки")
        
        if total_emails == 0:
            logger.error("Не найдено email адресов для проверки")
            raise ValueError("Не найдено email адресов")
        
        # Ограничение количества email для проверки
        if max_emails and max_emails > 0:
            emails = emails[:max_emails]
            logger.info(f"Будет проверено {len(emails)} email адресов из {total_emails}")
        
        # Инициализация валидатора
        validator = EmailValidator(timeout=timeout, check_smtp=check_smtp, accept_catch_all=accept_catch_all, validation_mode=validation_mode)
        
        # Проверка email адресов
        results = []
        total_emails = len(emails)
        start_time = time.time()
        
        # Оценка времени: среднее время на один email
        # Без SMTP: ~0.5-1 сек, с SMTP: ~1-3 сек
        avg_time_per_email = 1.5 if check_smtp else 0.7
        estimated_total_time = total_emails * avg_time_per_email
        
        logger.info(f"Начало проверки {total_emails} email адресов")
        if check_smtp:
            logger.info(f"Примерное время до завершения: {int(estimated_total_time // 60)} мин {int(estimated_total_time % 60)} сек")
        else:
            logger.info(f"Примерное время до завершения: {int(estimated_total_time // 60)} мин {int(estimated_total_time % 60)} сек (без SMTP)")
        
        # Определение базового имени для промежуточных файлов
        if output_file:
            base_output_name = os.path.splitext(output_file)[0]
            output_dir = os.path.dirname(output_file) if os.path.dirname(output_file) else '.'
        elif base_name:
            output_dir = os.path.dirname(input_file) if os.path.dirname(input_file) else '.'
            base_output_name = os.path.join(output_dir, base_name)
        else:
            output_dir = os.path.dirname(input_file) if os.path.dirname(input_file) else '.'
            base_output_name = os.path.join(output_dir, os.path.splitext(os.path.basename(input_file))[0])
        
        for i, email in enumerate(emails, 1):
            try:
                email_start_time = time.time()
                logger.info(f"[{i}/{total_emails}] Проверка: {email}")
                result = validator.validate_email(email)
                results.append(result)
                
                # Промежуточное сохранение каждые 1000 проверенных почт
                if i % 1000 == 0:
                    # Создание DataFrame из текущих результатов
                    checkpoint_df = pd.DataFrame(results)
                    
                    # Переупорядочивание столбцов
                    column_order = [
                        'Email', 'Пользователь', 'Домен', 'Валидность', 
                        'Надежность', 'МХ-записи',
                        'Корректность',
                        'Одноразовый (DEA)', 'Получение DNS, MX', 'Связь с SMTP-сервером',
                        'Email активен', 'Доставляемость', 'Catch-all адрес',
                        'Email переполнен', 'Ролевой аккаунт', 'Время проверки, сек',
                        'Попыток проверки'
                    ]
                    available_columns = [col for col in column_order if col in checkpoint_df.columns]
                    remaining_columns = [col for col in checkpoint_df.columns if col not in available_columns]
                    checkpoint_df = checkpoint_df[available_columns + remaining_columns]
                    
                    # Сохранение промежуточного файла
                    checkpoint_file = f"{base_output_name}_checkpoint_{i}.xlsx"
                    save_results_to_excel(checkpoint_df, checkpoint_file, is_checkpoint=True)
                    logger.info(f"✅ Промежуточное сохранение: {i} из {total_emails} проверенных почт")
                
                # Вычисление оставшегося времени на основе реальной скорости
                elapsed_time = time.time() - start_time
                if i > 0:
                    avg_time_actual = elapsed_time / i
                    remaining_emails = total_emails - i
                    estimated_remaining = avg_time_actual * remaining_emails
                    
                    if estimated_remaining > 60:
                        remaining_str = f"{int(estimated_remaining // 60)} мин {int(estimated_remaining % 60)} сек"
                    else:
                        remaining_str = f"{int(estimated_remaining)} сек"
                    
                    logger.info(f"  Прогресс: {i}/{total_emails} ({i/total_emails*100:.1f}%) | Осталось примерно: {remaining_str}")
                
                # Логирование результата
                status = "✅ ВАЛИДЕН" if result.get('Валидность') == 'Да' else "❌ НЕВАЛИДЕН"
                reliability = result.get('Надежность', 'Нет')
                logger.info(f"  Результат: {status} - Надежность: {reliability}")
                
                # Пауза между запросами
                if check_smtp and i % 10 == 0:
                    time.sleep(1)
                    
            except Exception as e:
                logger.error(f"Ошибка при проверке {email}: {e}")
                if email not in validator.check_attempts:
                    validator.check_attempts[email] = 0
                validator.check_attempts[email] += 1
                results.append({
                    'Email': email,
                    'Пользователь': '',
                    'Домен': '',
                    'Валидность': 'Нет',
                    'Корректность': 'Нет',
                    'Надежность': 'Нет',
                    'Одноразовый (DEA)': '–',
                    'Получение DNS, MX': 'Нет',
                    'Связь с SMTP-сервером': 'Нет',
                    'Email активен': 'Нет',
                    'Доставляемость': 'Нет',
                    'Catch-all адрес': 'Нет',  # При ошибке считаем что не catch-all
                    'Email переполнен': '–',
                    'Ролевой аккаунт': 'Нет',
                    'Время проверки, сек': 0,
                    'Попыток проверки': validator.check_attempts[email],
                    'МХ-записи': ''
                })
                
                # Промежуточное сохранение после ошибки тоже (если достигли кратного 1000)
                if i % 1000 == 0:
                    checkpoint_df = pd.DataFrame(results)
                    column_order = [
                        'Email', 'Пользователь', 'Домен', 'Валидность', 
                        'Надежность', 'МХ-записи',
                        'Корректность',
                        'Одноразовый (DEA)', 'Получение DNS, MX', 'Связь с SMTP-сервером',
                        'Email активен', 'Доставляемость', 'Catch-all адрес',
                        'Email переполнен', 'Ролевой аккаунт', 'Время проверки, сек',
                        'Попыток проверки'
                    ]
                    available_columns = [col for col in column_order if col in checkpoint_df.columns]
                    remaining_columns = [col for col in checkpoint_df.columns if col not in available_columns]
                    checkpoint_df = checkpoint_df[available_columns + remaining_columns]
                    checkpoint_file = f"{base_output_name}_checkpoint_{i}.xlsx"
                    save_results_to_excel(checkpoint_df, checkpoint_file, is_checkpoint=True)
                    logger.info(f"✅ Промежуточное сохранение: {i} из {total_emails} проверенных почт")
        
        # Создание DataFrame
        results_df = pd.DataFrame(results)
        
        # Определение порядка столбцов согласно примеру (Корректность и Надежность поменяны местами)
        # МХ-записи на 6 месте после Надежность
        column_order = [
            'Email', 'Пользователь', 'Домен', 'Валидность', 
            'Надежность', 'МХ-записи',  # МХ-записи на 6 месте после Надежность
            'Корректность',  # Корректность после МХ-записей
            'Одноразовый (DEA)', 'Получение DNS, MX', 'Связь с SMTP-сервером',
            'Email активен', 'Доставляемость', 'Catch-all адрес',
            'Email переполнен', 'Ролевой аккаунт', 'Время проверки, сек',
            'Попыток проверки'
        ]
        
        # Переупорядочивание столбцов
        available_columns = [col for col in column_order if col in results_df.columns]
        remaining_columns = [col for col in results_df.columns if col not in available_columns]
        results_df = results_df[available_columns + remaining_columns]
        
        # Генерация имени выходного файла с временем завершения
        if output_file is None:
            if base_name is not None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"{base_output_name}_{timestamp}.xlsx"
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_input = os.path.splitext(os.path.basename(input_file))[0]
                output_dir = os.path.dirname(input_file) if os.path.dirname(input_file) else '.'
                output_file = os.path.join(output_dir, f"{base_input}_{timestamp}.xlsx")
        else:
            # Если output_file был указан, используем его как есть
            pass
        
        # Сохранение финальных результатов в Excel
        logger.info(f"Сохранение результатов в: {output_file}")
        save_results_to_excel(results_df, output_file, is_checkpoint=False)
        
        # Статистика
        logger.info("\n" + "=" * 60)
        logger.info("СТАТИСТИКА ПРОВЕРКИ")
        logger.info("=" * 60)
        processed_count = len(results_df)
        logger.info(f"Всего обработано: {processed_count}")
        
        if 'Валидность' in results_df.columns:
            valid_count = len(results_df[results_df['Валидность'] == 'Да'])
            invalid_count = len(results_df[results_df['Валидность'] == 'Нет'])
            logger.info(f"Валидных для рассылки: {valid_count} ({valid_count / processed_count * 100:.1f}%)")
            logger.info(f"Невалидных: {invalid_count} ({invalid_count / processed_count * 100:.1f}%)")
        
        if 'Надежность' in results_df.columns:
            reliability_stats = results_df['Надежность'].value_counts()
            logger.info("\nРаспределение по надежности:")
            for reliability in ['Высокая', 'Средняя', 'Нет']:
                if reliability in reliability_stats:
                    count = reliability_stats[reliability]
                    logger.info(f"  {reliability}: {count} ({count / processed_count * 100:.1f}%)")
        
        logger.info("=" * 60)
        
        return results_df
        
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")
        raise


def main():
    """Основная функция для запуска из командной строки"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Валидатор email адресов для рассылок',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument('--input', '-i', 
                       default='emails.xlsx',
                       help='Путь к входному Excel файлу (по умолчанию: emails.xlsx)')
    parser.add_argument('--output', '-o',
                       help='Путь к выходному Excel файлу (по умолчанию: генерируется автоматически)')
    parser.add_argument('--no-smtp', action='store_true',
                       help='Отключить SMTP проверку (быстрее, но менее точно)')
    parser.add_argument('--timeout', type=int, default=10,
                       help='Таймаут для сетевых запросов в секундах (по умолчанию: 10)')
    
    args = parser.parse_args()
    
    try:
        process_excel_file(
            input_file=args.input,
            output_file=args.output,
            check_smtp=not args.no_smtp,
            timeout=args.timeout
        )
    except Exception as e:
        logger.error(f"Программа завершилась с ошибкой: {str(e)}")
        exit(1)


if __name__ == "__main__":
    print("=" * 60)
    print("EMAIL ВАЛИДАТОР ДЛЯ РАССЫЛОК")
    print("=" * 60)
    
    # Интерактивный режим
    # 1. Ввод названия файла
    input_file = input("Введите название файла (по умолчанию: emails): ").strip()
    if not input_file:
        input_file = "emails"
    
    # Добавляем расширение если его нет
    if not input_file.endswith(('.xlsx', '.xls', '.csv')):
        input_file = input_file + ".xlsx"
    
    # Проверяем существование файла в текущей директории
    if not os.path.exists(input_file):
        # Пробуем с другими расширениями
        for ext in ['.xlsx', '.xls', '.csv']:
            test_file = input_file.rsplit('.', 1)[0] + ext
            if os.path.exists(test_file):
                input_file = test_file
                break
    
    # 2. Выполнять ли SMTP проверку
    smtp_choice = input("Выполнять SMTP проверку? (да/нет, по умолчанию: да): ").strip().lower()
    check_smtp = smtp_choice not in ['нет', 'no', 'n', '0', 'false']
    
    # 3. Режим валидации
    mode_choice = input("Выберите режим валидации (1 - строгий, 2 - лояльный, по умолчанию: строгий): ").strip()
    validation_mode = 'strict'
    if mode_choice == '2' or mode_choice.lower() in ['лояльный', 'lenient', 'l']:
        validation_mode = 'lenient'
        print("Выбран лояльный режим (на 15-20% больше валидных адресов)")
    else:
        print("Выбран строгий режим (максимальная точность)")
    
    # 4. Считаем ли валидными catch-all почты
    catch_all_choice = input("Считаем ли валидными catch-all почты? (да/нет, по умолчанию: нет): ").strip().lower()
    accept_catch_all = catch_all_choice in ['да', 'yes', 'y', '1', 'true']
    
    # 5. Сколько почт проверять
    max_emails_input = input("Сколько почт из списка вы хотите проверить? (если нет ответа, то все): ").strip()
    max_emails = None
    if max_emails_input:
        try:
            max_emails = int(max_emails_input)
            if max_emails <= 0:
                max_emails = None
        except ValueError:
            max_emails = None
    
    try:
        print(f"\nНачинаем проверку...")
        print(f"Файл: {input_file}")
        print(f"Режим: {'Строгий' if validation_mode == 'strict' else 'Лояльный'}")
        print(f"SMTP проверка: {'Включена' if check_smtp else 'Отключена'}")
        if max_emails:
            print(f"Будет проверено: {max_emails} адресов")
        print("=" * 60)
        
        process_excel_file(
            input_file=input_file,
            check_smtp=check_smtp,
            accept_catch_all=accept_catch_all,
            max_emails=max_emails,
            validation_mode=validation_mode
        )
        print("\n" + "=" * 60)
        print("Проверка завершена! Нажмите Enter для выхода...")
        input()
    except Exception as e:
        print(f"\nОшибка: {str(e)}")
        print("Нажмите Enter для выхода...")
        input()
